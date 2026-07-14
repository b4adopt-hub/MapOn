# -*- coding: utf-8 -*-
"""LURIS(토지이음) 월별 파일 -> Supabase 적재
zoning_rates: 시군구x용도지역 건폐율(bcr)/용적률(far), 기본율/특례 분류, 법정상한 검수 플래그
  - 기본율 판정: (1) 시군구x구분별 self_match 용도지역 수 최대 조항 → (2) 그 조항 내 최빈 항
    (방화지구 완화·성장관리구역 등 특례가 같은 조의 별도 항에 붙는 구조를 배제)
permitted_uses: 시군구x용도지역지구x토지이용행위 가능여부
같은 --month 재실행 시 해당 월 데이터를 삭제 후 재적재 (멱등)
"""
import argparse, io, os, re, sys, csv
from collections import defaultdict, Counter

from openpyxl import load_workbook
import psycopg2

PCT = re.compile(r'(\d+(?:\.\d+)?)\s*퍼센트')
PCT100 = re.compile(r'100분의\s*(\d+(?:\.\d+)?)')
ITEM = re.compile(r'^\s*\d+\.\s*(?:제?\d*종?\s*)?[가-힣·ㆍ0-9\s]+(지역|지구)\s*:\s*(\d+(?:\.\d+)?)\s*퍼센트\s*이하')
JO = re.compile(r'(제\d+조(?:의\d+)?)')
HANG = re.compile(r'제(\d+)항')

BCR_CAP = {'UQA111':50,'UQA112':50,'UQA121':60,'UQA122':60,'UQA123':50,'UQA130':70,'UQA210':90,'UQA220':80,'UQA230':70,'UQA240':80,'UQA310':70,'UQA320':70,'UQA330':70,'UQA410':20,'UQA420':20,'UQA430':20,'UQB100':40,'UQB200':20,'UQB300':20,'UQC001':20,'UQD001':20}
FAR_CAP = {'UQA111':100,'UQA112':150,'UQA121':200,'UQA122':250,'UQA123':300,'UQA130':500,'UQA210':1500,'UQA220':1300,'UQA230':900,'UQA240':1100,'UQA310':300,'UQA320':350,'UQA330':400,'UQA410':80,'UQA420':100,'UQA430':100,'UQB100':100,'UQB200':80,'UQB300':80,'UQC001':80,'UQD001':80}

def norm(s):
    return re.sub(r'[\s·ㆍ()]', '', s or '')

def parse_law(path):
    ws = load_workbook(path, read_only=True).worksheets[0]
    raw = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        code = str(r[4] or '').strip()
        if not code or code == '00000':
            continue
        is_bp, is_yr = str(r[9]).strip() == '○', str(r[10]).strip() == '○'
        if not (is_bp or is_yr):
            continue
        content = str(r[11] or '')
        prov = str(r[7] or '')
        jo_m = JO.search(prov)
        hang_m = HANG.search(prov)
        m = ITEM.match(content.strip())
        zone_in = norm(content.strip().split(':')[0].split('.', 1)[-1]) if m else None
        self_match = bool(m) and zone_in and (zone_in in norm(str(r[2])) or norm(str(r[2])) in zone_in)
        vals = [float(x) for x in PCT.findall(content)] + [float(x) for x in PCT100.findall(content)]
        for kind, flag in (('bcr', is_bp), ('far', is_yr)):
            if flag:
                raw.append(dict(code=code, zone_cd=str(r[1]).strip(), zone_nm=str(r[2]).strip(),
                                kind=kind, jo=jo_m.group(1) if jo_m else '',
                                hang=hang_m.group(1) if hang_m else '0', prov=prov.strip(),
                                ord_nm=str(r[5]).strip(), dt=str(r[6]).strip(), vals=vals,
                                self_match=self_match, item_val=float(m.group(2)) if m else None,
                                content=content))
    # (1) 기본율 조항 = self_match 용도지역 수 최대 조번호
    grp = defaultdict(set)
    for x in raw:
        if x['self_match']:
            grp[(x['code'], x['kind'], x['jo'])].add(x['zone_cd'])
    base_jo = {}
    for (code, kind, jo), zones in grp.items():
        k = (code, kind)
        if k not in base_jo or len(zones) > len(grp[(code, kind, base_jo[k])]):
            base_jo[k] = jo
    # (2) 그 조항 내 최빈 항 = 기본율 항 (특례 항 배제)
    hang_cnt = defaultdict(Counter)
    for x in raw:
        if x['self_match'] and x['jo'] == base_jo.get((x['code'], x['kind'])):
            hang_cnt[(x['code'], x['kind'])][x['hang']] += 1
    base_hang = {k: c.most_common(1)[0][0] for k, c in hang_cnt.items()}

    out = []
    for x in raw:
        k = (x['code'], x['kind'])
        is_base = x['self_match'] and x['jo'] == base_jo.get(k) and x['hang'] == base_hang.get(k)
        cap = (BCR_CAP if x['kind'] == 'bcr' else FAR_CAP).get(x['zone_cd'])
        review = bool(is_base and cap and x['item_val'] and x['item_val'] > cap)
        dt = x['dt']
        out.append((x['code'], x['zone_cd'], x['zone_nm'], x['kind'],
                    'base' if is_base else 'special',
                    x['item_val'] if is_base else None,
                    ';'.join(str(v) for v in x['vals'])[:200],
                    x['ord_nm'][:200], x['prov'][:100],
                    f"{dt[:4]}-{dt[4:6]}-{dt[6:8]}" if len(dt) == 8 and dt.isdigit() else None,
                    x['content'][:2000], review))
    return out

def parse_act(path):
    wb = load_workbook(path, read_only=True)
    for sheet in wb.sheetnames:
        is_ord = sheet != '법령'
        ws = wb[sheet]
        for r in ws.iter_rows(min_row=2, values_only=True):
            code = str(r[0] or '').strip()
            land_use = str(r[4] or '').strip()
            if not code or not land_use:
                continue
            yield (code, str(r[1] or '').strip(), str(r[2] or '').strip(),
                   str(r[3] or '').strip()[:300], land_use[:300],
                   str(r[5] or '').strip()[:200], str(r[6] or '').strip()[:2000], is_ord)

def copy_rows(cur, table, cols, rows, month, batch=50000):
    buf = io.StringIO()
    w = csv.writer(buf)
    n = 0
    def flush():
        nonlocal buf, w
        buf.seek(0)
        cur.copy_expert(f"copy {table} ({','.join(cols)},src_month) from stdin with (format csv)", buf)
        buf = io.StringIO()
        w = csv.writer(buf)
    for row in rows:
        w.writerow([('' if v is None else v) for v in row] + [month])
        n += 1
        if n % batch == 0:
            flush()
            print(f'  {table}: {n:,}행')
    flush()
    print(f'  {table}: 총 {n:,}행 완료')

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--law', required=True)
    ap.add_argument('--act', required=True)
    ap.add_argument('--month', required=True)
    args = ap.parse_args()
    url = os.environ.get('DATABASE_URL')
    if not url:
        sys.exit('DATABASE_URL 환경변수 필요')

    conn = psycopg2.connect(url)
    conn.autocommit = False
    cur = conn.cursor()

    print('[1/2] zoning_rates')
    zr = parse_law(args.law)
    cur.execute('delete from public.zoning_rates where src_month = %s', (args.month,))
    copy_rows(cur, 'public.zoning_rates',
              ['sgg_code','zone_cd','zone_nm','rate_kind','category','rate_pct','rate_values','ordinance','provision','enforce_dt','content','needs_review'],
              zr, args.month)

    print('[2/2] permitted_uses (약 330만 행)')
    cur.execute('delete from public.permitted_uses where src_month = %s', (args.month,))
    copy_rows(cur, 'public.permitted_uses',
              ['sgg_code','sgg_name','zone_nm','law_name','land_use','decision','condition_note','is_ordinance'],
              parse_act(args.act), args.month)

    conn.commit()
    cur.execute('select category, count(*) from public.zoning_rates group by 1 order by 1')
    print('zoning_rates:', cur.fetchall())
    cur.execute('select count(*) from public.zoning_rates where needs_review')
    print('검수 대상(needs_review):', cur.fetchone()[0])
    cur.execute('select count(*) from public.permitted_uses')
    print('permitted_uses:', cur.fetchone()[0])
    conn.close()
    print('완료')

if __name__ == '__main__':
    main()
